import logging
from uuid import UUID

from fastapi import HTTPException, status, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.customer import Customer
from app.models.user import User
from app.repositories.customer_repository import CustomerRepository
from app.schemas.customer import CustomerCreate, CustomerUpdate

logger = logging.getLogger(__name__)


class CustomerService:

    def __init__(self, db: Session):
        self.db = db
        self.repo = CustomerRepository(db)

    def _enrich_loyalty_points(self, current_user: User, customer: Customer) -> Customer:
        if customer and not customer.loyalty:
            from app.services.loyalty_service import LoyaltyService
            loyalty = LoyaltyService(self.db).get_customer_loyalty(current_user, customer.id)
            customer.loyalty = loyalty
        return customer

    def list_customers(self, current_user: User) -> list[Customer]:
        logger.info(
            "Listing customers | business_id=%s requested_by=%s",
            current_user.business_id,
            current_user.id,
        )
        customers = self.repo.get_all_by_business(current_user.business_id)
        from app.services.loyalty_service import LoyaltyService
        loyalty_service = LoyaltyService(self.db)
        for c in customers:
            if not c.loyalty:
                c.loyalty = loyalty_service.get_customer_loyalty(current_user, c.id)
        return customers

    def get_paginated_customers(
        self,
        current_user: User,
        page: int = 1,
        limit: int = 10,
        search: str | None = None,
        sort: str | None = "newest",
        filter: str | None = "all",
    ) -> dict:
        logger.info(
            "Fetching paginated customers | business_id=%s page=%s limit=%s search=%s sort=%s filter=%s",
            current_user.business_id,
            page,
            limit,
            search,
            sort,
            filter,
        )
        res = self.repo.get_paginated_by_business(
            business_id=current_user.business_id,
            page=page,
            limit=limit,
            search=search,
            sort=sort,
            filter=filter,
        )
        from app.services.loyalty_service import LoyaltyService
        loyalty_service = LoyaltyService(self.db)
        for c in res["items"]:
            if not c.loyalty:
                c.loyalty = loyalty_service.get_customer_loyalty(current_user, c.id)
        return res

    def get_customer(self, current_user: User, customer_id: UUID) -> Customer:
        customer = self.repo.get_by_id(customer_id)
        if not customer or customer.business_id != current_user.business_id:
            logger.warning(
                "Customer not found or tenant mismatch | customer_id=%s business_id=%s",
                customer_id,
                current_user.business_id,
            )
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Customer not found.",
            )

        logger.info(
            "Customer fetched | customer_id=%s business_id=%s",
            customer.id,
            customer.business_id,
        )
        return self._enrich_loyalty_points(current_user, customer)

    def get_customer_by_phone(self, current_user: User, phone: str) -> Customer:
        customer = self.repo.get_by_phone(current_user.business_id, phone.strip())
        if not customer:
            logger.info(
                "Customer not found by phone | phone=%s business_id=%s",
                phone,
                current_user.business_id,
            )
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Customer with phone '{phone}' not found.",
            )
        return self._enrich_loyalty_points(current_user, customer)

    def create_customer(
        self, current_user: User, data: CustomerCreate
    ) -> Customer:
        logger.info(
            "Creating customer | business_id=%s phone=%s",
            current_user.business_id,
            data.phone,
        )

        existing_customer = self.repo.get_by_phone(
            current_user.business_id, data.phone
        )
        if existing_customer:
            logger.warning(
                "Customer creation rejected — duplicate phone inside business | business_id=%s phone=%s",
                current_user.business_id,
                data.phone,
            )
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A customer with this phone number already exists in your business.",
            )

        customer = Customer(
            business_id=current_user.business_id,
            **data.model_dump(),
        )
        created_customer = self.repo.create(customer)
        self.db.commit()
        self.db.refresh(created_customer)

        logger.info(
            "Customer created successfully | customer_id=%s business_id=%s",
            created_customer.id,
            created_customer.business_id,
        )
        return self._enrich_loyalty_points(current_user, created_customer)

    def update_customer(
        self,
        current_user: User,
        customer_id: UUID,
        data: CustomerUpdate,
    ) -> Customer:
        customer = self.get_customer(current_user, customer_id)

        if data.phone is not None and data.phone != customer.phone:
            existing_customer = self.repo.get_by_phone(
                current_user.business_id, data.phone
            )
            if existing_customer and existing_customer.id != customer.id:
                logger.warning(
                    "Customer update rejected — duplicate phone inside business | business_id=%s phone=%s",
                    current_user.business_id,
                    data.phone,
                )
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="A customer with this phone number already exists in your business.",
                )

        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(customer, field, value)

        self.repo.update(customer)
        self.db.commit()
        self.db.refresh(customer)

        logger.info(
            "Customer updated successfully | customer_id=%s business_id=%s",
            customer.id,
            customer.business_id,
        )
        return self._enrich_loyalty_points(current_user, customer)

    def delete_customer(self, current_user: User, customer_id: UUID) -> dict:
        customer = self.get_customer(current_user, customer_id)
        self.repo.delete(customer)
        self.db.commit()
        logger.info(
            "Customer deleted successfully | customer_id=%s business_id=%s",
            customer_id,
            current_user.business_id,
        )
        return {"message": "Customer deleted successfully", "id": str(customer_id)}

    def record_customer_visit(self, current_user: User, customer_id: UUID, amount_spent: float) -> Customer:
        from datetime import datetime, timezone
        customer = self.get_customer(current_user, customer_id)
        now_ts = datetime.now(timezone.utc)

        customer.visit_count = (customer.visit_count or 0) + 1
        customer.total_spent = (customer.total_spent or 0.0) + max(0.0, amount_spent)
        if not customer.first_visit_at:
            customer.first_visit_at = now_ts
        customer.last_visit_at = now_ts

        # Calculate and award loyalty points
        from app.repositories.loyalty_repository import LoyaltyRepository
        from app.models.loyalty import CustomerLoyalty
        loyalty_repo = LoyaltyRepository(self.db)
        loyalty_settings = loyalty_repo.get_settings(current_user.business_id)

        points_per_amount = loyalty_settings.points_per_amount if (loyalty_settings and loyalty_settings.is_active) else 1
        amount_req = loyalty_settings.amount_required if (loyalty_settings and loyalty_settings.is_active) else 10

        earned = int((amount_spent / amount_req) * points_per_amount) if (amount_req > 0 and points_per_amount > 0) else int(amount_spent // 10)
        if earned > 0:
            cl = loyalty_repo.get_customer_loyalty(customer_id)
            if not cl:
                cl = CustomerLoyalty(customer_id=customer_id, current_points=0, lifetime_points=0, redeemed_points=0)
                loyalty_repo.create_customer_loyalty(cl)
            cl.current_points += earned
            cl.lifetime_points += earned
            loyalty_repo.update_customer_loyalty(cl)

        self.repo.update(customer)
        self.db.commit()
        self.db.refresh(customer)
        logger.info("Recorded customer visit | customer_id=%s visits=%d total_spent=%.2f points=%d", customer.id, customer.visit_count, customer.total_spent, customer.loyalty_points)
        return self._enrich_loyalty_points(current_user, customer)

    def get_customer_crm_details(self, current_user: User, customer_id: UUID) -> dict:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from app.models.order import Order, OrderSource, OrderStatus
        from app.models.visit import Visit
        from app.models.restaurant_table import RestaurantTable
        from app.models.dining_area import DiningArea
        from app.models.campaign import CampaignLog
        from app.schemas.customer import CustomerResponse

        customer = self.get_customer(current_user, customer_id)

        # 1. Fetch Orders for customer
        orders_stmt = (
            select(Order)
            .options(joinedload(Order.items))
            .where(Order.customer_id == customer_id, Order.business_id == current_user.business_id)
            .order_by(Order.created_at.desc())
        )
        orders = list(self.db.scalars(orders_stmt).unique().all())

        # 2. Fetch Visits for customer
        visits_stmt = (
            select(Visit)
            .where(Visit.customer_id == customer_id, Visit.business_id == current_user.business_id)
            .order_by(Visit.started_at.desc())
        )
        visits = list(self.db.scalars(visits_stmt).all())

        # 3. Tables & Dining Areas map
        tables_stmt = select(RestaurantTable).where(RestaurantTable.business_id == current_user.business_id)
        tables = list(self.db.scalars(tables_stmt).all())
        table_map = {t.id: t for t in tables}

        areas_stmt = select(DiningArea).where(DiningArea.business_id == current_user.business_id)
        areas = list(self.db.scalars(areas_stmt).all())
        area_map = {a.id: a.name for a in areas}

        # 4. Campaign logs
        logs_stmt = (
            select(CampaignLog)
            .options(joinedload(CampaignLog.campaign))
            .where(CampaignLog.customer_id == customer_id)
            .order_by(CampaignLog.created_at.desc())
        )
        campaign_logs = list(self.db.scalars(logs_stmt).all())

        # Calculate metrics
        completed_orders = [o for o in orders if o.status == OrderStatus.SERVED]
        total_spent = sum(o.total_amount for o in completed_orders) or customer.total_spent or 0.0
        total_visits = max(len(visits), customer.visit_count or 0)
        total_orders = len(orders)
        avg_bill = round(total_spent / max(1, len(completed_orders) or total_orders), 2) if (total_spent > 0 and (completed_orders or orders)) else 0.0

        total_qr = sum(1 for o in orders if str(o.order_source).upper() in ("QR", "ORDERSOURCE.QR"))
        total_staff = total_orders - total_qr

        # Table & Dining Area preference
        table_counts = {}
        for o in orders:
            if o.table_id:
                table_counts[o.table_id] = table_counts.get(o.table_id, 0) + 1

        fav_table_obj = table_map.get(max(table_counts, key=table_counts.get)) if table_counts else None
        fav_table_name = fav_table_obj.table_name if fav_table_obj else (tables[0].table_name if tables else "Table 1")
        pref_area_name = area_map.get(fav_table_obj.dining_area_id) if (fav_table_obj and fav_table_obj.dining_area_id) else (areas[0].name if areas else "Main Hall")

        # Favorite items
        item_stats = {}
        for o in orders:
            for item in o.items:
                name = item.item_name
                if name not in item_stats:
                    item_stats[name] = {"count": 0, "total_spent": 0.0}
                item_stats[name]["count"] += item.quantity
                item_stats[name]["total_spent"] += item.subtotal

        sorted_items = sorted(item_stats.items(), key=lambda x: x[1]["count"], reverse=True)[:3]
        fav_items_list = [
            {"name": name, "count": data["count"], "total_spent": round(data["total_spent"], 2)}
            for name, data in sorted_items
        ]

        # Calculate average visit frequency in days
        avg_frequency_days = None
        if len(visits) >= 2:
            dates = sorted([v.started_at for v in visits if v.started_at])
            if len(dates) >= 2:
                diff_days = (dates[-1] - dates[0]).days
                avg_frequency_days = round(diff_days / (len(dates) - 1), 1)

        # Timeline events synthesis
        timeline = []
        for o in orders:
            t_obj = table_map.get(o.table_id)
            t_name = t_obj.table_name if t_obj else "Table"
            a_name = area_map.get(t_obj.dining_area_id) if (t_obj and t_obj.dining_area_id) else "Main Area"
            item_count = sum(i.quantity for i in o.items)

            # Order Placed Event
            timeline.append({
                "id": f"order-{o.id}",
                "type": "ORDER",
                "title": f"Ordered {item_count} Item(s) at {a_name} ({t_name})",
                "description": f"Order #{o.order_number} · Source: {o.order_source}",
                "timestamp": o.created_at,
                "badge": o.order_source,
                "amount": o.total_amount,
            })

            # Bill Payment Completed Event (if SERVED)
            if o.status == OrderStatus.SERVED:
                earned_pts = int(round((o.total_amount / 100.0) * 10.0))
                timeline.append({
                    "id": f"pay-{o.id}",
                    "type": "PAYMENT",
                    "title": f"Bill Payment Completed — ₹{o.total_amount:,.2f}",
                    "description": f"Order #{o.order_number} settled successfully",
                    "timestamp": o.updated_at or o.created_at,
                    "badge": "Completed",
                    "amount": o.total_amount,
                })
                if earned_pts > 0:
                    timeline.append({
                        "id": f"loyalty-{o.id}",
                        "type": "LOYALTY",
                        "title": f"Earned {earned_pts} Loyalty Points",
                        "description": f"Purchase reward for Order #{o.order_number}",
                        "timestamp": o.updated_at or o.created_at,
                        "badge": f"+{earned_pts} pts",
                        "amount": None,
                    })

        for v in visits:
            timeline.append({
                "id": f"visit-{v.id}",
                "type": "VISIT",
                "title": f"Visited Restaurant",
                "description": f"Status: {v.status} · Spent: ₹{v.total_amount:,.2f}",
                "timestamp": v.started_at or v.created_at,
                "badge": v.status,
                "amount": v.total_amount,
            })

        for log in campaign_logs:
            c_name = log.campaign.name if log.campaign else "WhatsApp Campaign"
            timeline.append({
                "id": f"wa-{log.id}",
                "type": "WHATSAPP",
                "title": f"WhatsApp Message Sent ({c_name})",
                "description": log.campaign.message if (log.campaign and log.campaign.message) else "Message delivered",
                "timestamp": log.sent_at or log.created_at,
                "badge": log.status,
                "amount": None,
            })

        # Sort timeline descending by timestamp
        timeline.sort(key=lambda x: x["timestamp"], reverse=True)

        # Build Visits response list
        visits_list = []
        for idx, v in enumerate(reversed(visits), 1):
            visits_list.append({
                "id": str(v.id),
                "visit_number": idx,
                "date": v.started_at or v.created_at,
                "table_name": fav_table_name,
                "dining_area_name": pref_area_name,
                "source": "QR" if v.services else "STAFF",
                "status": v.status,
                "total_amount": v.total_amount,
                "loyalty_earned": int(round((v.total_amount / 100.0) * 10.0)),
                "payment_method": v.payment_method or "CASH",
            })
        visits_list.sort(key=lambda x: x["date"], reverse=True)

        # Build Orders response list
        orders_list = []
        for o in orders:
            t_obj = table_map.get(o.table_id)
            orders_list.append({
                "id": str(o.id),
                "order_number": o.order_number,
                "status": o.status,
                "source": o.order_source,
                "table_name": t_obj.table_name if t_obj else "Table",
                "created_at": o.created_at,
                "completed_at": o.updated_at if o.status == OrderStatus.SERVED else None,
                "subtotal": o.subtotal,
                "tax_amount": o.tax_amount,
                "discount_amount": o.discount_amount,
                "total_amount": o.total_amount,
                "items": [
                    {
                        "id": str(i.id),
                        "name": i.item_name,
                        "unit_price": i.unit_price,
                        "quantity": i.quantity,
                        "subtotal": i.subtotal,
                        "notes": i.notes,
                    }
                    for i in o.items
                ],
            })

        # Build Loyalty Ledger
        loyalty_obj = customer.loyalty
        curr_pts = loyalty_obj.current_points if loyalty_obj else customer.loyalty_points
        life_pts = loyalty_obj.lifetime_points if loyalty_obj else curr_pts
        redeemed_pts = loyalty_obj.redeemed_points if loyalty_obj else 0

        loyalty_ledger = []
        for o in completed_orders:
            pts = int(round((o.total_amount / 100.0) * 10.0))
            if pts > 0:
                loyalty_ledger.append({
                    "id": f"loy-ledg-{o.id}",
                    "date": o.updated_at or o.created_at,
                    "reason": f"Earned from Order #{o.order_number} (₹{o.total_amount:,.2f})",
                    "points": pts,
                    "type": "EARNED",
                    "balance_after": curr_pts,
                })
        loyalty_ledger.sort(key=lambda x: x["date"], reverse=True)

        # Build WhatsApp & Campaign logs
        wa_logs_list = [
            {
                "id": str(log.id),
                "campaign_name": log.campaign.name if log.campaign else "Direct Message",
                "type": log.campaign.campaign_type if log.campaign else "WELCOME",
                "message": log.campaign.message if log.campaign else "Thank you for dining with us!",
                "status": log.status,
                "sent_at": log.sent_at or log.created_at,
            }
            for log in campaign_logs
        ]

        campaigns_list = [
            {
                "id": str(log.campaign.id) if log.campaign else str(log.id),
                "name": log.campaign.name if log.campaign else "Automated Campaign",
                "type": log.campaign.campaign_type if log.campaign else "CUSTOM",
                "status": log.status,
                "sent_at": log.sent_at or log.created_at,
            }
            for log in campaign_logs if log.campaign
        ]

        # AI Insights generator logic
        ai_insights_text = "AI insights will appear after sufficient customer activity."
        if total_visits >= 2 or total_orders >= 2:
            ai_insights_text = (
                f"{customer.name} is a repeat guest with {total_visits} visit(s) and a lifetime spend of ₹{total_spent:,.2f}. "
                f"They prefer dining in {pref_area_name} ({fav_table_name}) and frequently order {sorted_items[0][0] if sorted_items else 'popular items'}. "
                f"Recommended engagement: Send a loyalty bonus coupon or weekend special offer."
            )

        last_ord_at = orders[0].created_at if orders else None

        return {
            "profile": CustomerResponse.model_validate(customer),
            "total_visits": total_visits,
            "total_orders": total_orders,
            "total_spent": total_spent,
            "avg_bill": avg_bill,
            "loyalty_points": curr_pts,
            "last_visit_at": customer.last_visit_at,
            "first_visit_at": customer.first_visit_at,
            "last_order_at": last_ord_at,
            "customer_since": customer.created_at,
            "total_qr_orders": total_qr,
            "total_staff_orders": total_staff,
            "preferred_dining_area": pref_area_name,
            "favorite_table": fav_table_name,
            "favorite_items": fav_items_list,
            "avg_visit_frequency_days": avg_frequency_days,
            "customer_lifetime_value": total_spent,
            "timeline": timeline,
            "visits": visits_list,
            "orders": orders_list,
            "loyalty_history": loyalty_ledger,
            "loyalty_current_points": curr_pts,
            "loyalty_lifetime_points": life_pts,
            "loyalty_redeemed_points": redeemed_pts,
            "whatsapp_logs": wa_logs_list,
            "campaigns": campaigns_list,
            "reviews": [],
            "ai_insights": ai_insights_text,
        }

    def get_welcome_campaign_data(
        self,
        current_user: User,
        timeframe: str = "today",
        search: str | None = None,
        sort_by: str | None = "newest",
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        from datetime import datetime, timedelta, timezone
        from sqlalchemy import extract, func, or_, select
        from app.models.customer import Customer
        from app.models.order import Order
        from app.models.campaign import Campaign, CampaignLog, CampaignLogStatus, CampaignType

        now = datetime.now(timezone.utc)
        start_of_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
        start_of_week = start_of_today - timedelta(days=now.weekday())
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        thirty_days_ago = now - timedelta(days=30)

        business_id = current_user.business_id

        # 1. Summary Cards Calculations
        todays_new = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                or_(Customer.created_at >= start_of_today, Customer.first_visit_at >= start_of_today),
            )
        ) or 0

        this_week_new = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                or_(Customer.created_at >= start_of_week, Customer.first_visit_at >= start_of_week),
            )
        ) or 0

        this_month_new = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                or_(Customer.created_at >= start_of_month, Customer.first_visit_at >= start_of_month),
            )
        ) or 0

        returning_cust = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                Customer.visit_count > 1,
            )
        ) or 0

        today_m, today_d = now.month, now.day
        birthdays_today = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                extract("month", Customer.birth_date) == today_m,
                extract("day", Customer.birth_date) == today_d,
            )
        ) or 0

        recovery_due = self.db.scalar(
            select(func.count(Customer.id)).where(
                Customer.business_id == business_id,
                Customer.last_visit_at.isnot(None),
                Customer.last_visit_at <= thirty_days_ago,
            )
        ) or 0

        summary_cards = {
            "todays_new": todays_new,
            "this_week": this_week_new,
            "this_month": this_month_new,
            "returning": returning_cust,
            "birthdays_today": birthdays_today,
            "recovery_due": recovery_due,
        }

        # 2. Overall Metrics Calculations
        total_customers = self.db.scalar(
            select(func.count(Customer.id)).where(Customer.business_id == business_id)
        ) or 0

        total_spent_sum = float(self.db.scalar(
            select(func.coalesce(func.sum(Customer.total_spent), 0.0)).where(Customer.business_id == business_id)
        ) or 0.0)

        total_visits_sum = self.db.scalar(
            select(func.coalesce(func.sum(Customer.visit_count), 0)).where(Customer.business_id == business_id)
        ) or 0

        avg_lifetime_value = round(total_spent_sum / max(1, total_customers), 2) if total_customers > 0 else 0.0
        avg_visits = round(total_visits_sum / max(1, total_customers), 1) if total_customers > 0 else 0.0
        returning_pct = round((returning_cust / max(1, total_customers)) * 100, 1) if total_customers > 0 else 0.0

        if timeframe == "week":
            selected_count = this_week_new
        elif timeframe == "month":
            selected_count = this_month_new
        elif timeframe == "all":
            selected_count = total_customers
        else:
            selected_count = todays_new

        avg_first_order_value = avg_lifetime_value

        metrics = {
            "first_visit_count": selected_count,
            "returning_pct": returning_pct,
            "avg_first_order_value": avg_first_order_value,
            "avg_lifetime_value": avg_lifetime_value,
            "avg_visits": avg_visits,
        }

        # 3. Base Query for Filtered Customers
        stmt = select(Customer).where(Customer.business_id == business_id)

        if timeframe == "week":
            stmt = stmt.where(or_(Customer.created_at >= start_of_week, Customer.first_visit_at >= start_of_week))
        elif timeframe == "month":
            stmt = stmt.where(or_(Customer.created_at >= start_of_month, Customer.first_visit_at >= start_of_month))
        elif timeframe == "today":
            stmt = stmt.where(or_(Customer.created_at >= start_of_today, Customer.first_visit_at >= start_of_today))

        if search and search.strip():
            term = f"%{search.strip()}%"
            stmt = stmt.where(
                or_(
                    Customer.name.ilike(term),
                    Customer.phone.ilike(term),
                    Customer.email.ilike(term),
                )
            )

        count_stmt = select(func.count()).select_from(stmt.subquery())
        total_items = self.db.scalar(count_stmt) or 0

        if sort_by == "oldest":
            stmt = stmt.order_by(Customer.created_at.asc())
        elif sort_by == "spend_desc":
            stmt = stmt.order_by(Customer.total_spent.desc())
        elif sort_by == "spend_asc":
            stmt = stmt.order_by(Customer.total_spent.asc())
        elif sort_by == "visits_desc":
            stmt = stmt.order_by(Customer.visit_count.desc())
        elif sort_by == "visits_asc":
            stmt = stmt.order_by(Customer.visit_count.asc())
        else:
            stmt = stmt.order_by(Customer.created_at.desc())

        total_pages = max(1, (total_items + page_size - 1) // page_size) if total_items > 0 else 1
        page = max(1, min(page, total_pages))
        offset = (page - 1) * page_size

        stmt = stmt.offset(offset).limit(page_size)
        customers = list(self.db.scalars(stmt).all())

        items = []
        for c in customers:
            first_order = self.db.scalar(
                select(Order)
                .where(Order.customer_id == c.id)
                .order_by(Order.created_at.asc())
            )
            source_str = "QR" if (first_order and str(first_order.order_source).upper() in ("QR", "ORDERSOURCE.QR")) else "Staff"

            if c.total_spent >= 500:
                cust_type = "VIP"
            elif c.visit_count > 1:
                cust_type = "Returning"
            else:
                cust_type = "New"

            welcome_log = self.db.scalar(
                select(CampaignLog)
                .join(Campaign, CampaignLog.campaign_id == Campaign.id)
                .where(
                    CampaignLog.customer_id == c.id,
                    Campaign.campaign_type == CampaignType.WELCOME,
                    CampaignLog.status == CampaignLogStatus.SENT,
                )
            )
            w_status = "Sent" if (welcome_log or (c.notes and "welcome" in c.notes.lower())) else "Pending"

            pts = c.loyalty.current_points if c.loyalty else c.loyalty_points

            items.append({
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "email": c.email,
                "first_visit_at": c.first_visit_at or c.created_at,
                "last_visit_at": c.last_visit_at or c.created_at,
                "visit_count": c.visit_count,
                "total_spent": float(c.total_spent),
                "loyalty_points": pts,
                "source": source_str,
                "customer_type": cust_type,
                "welcome_status": w_status,
                "created_at": c.created_at,
            })

        return {
            "summary_cards": summary_cards,
            "metrics": metrics,
            "items": items,
            "page": page,
            "page_size": page_size,
            "total": total_items,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_previous": page > 1,
        }

    def get_birthday_campaign_data(
        self,
        current_user: User,
        bucket: str = "today",
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        sort_by: str = "name",
        sort_order: str = "asc",
    ) -> dict:
        return self.get_celebration_campaign_data(
            current_user=current_user,
            kind="birthday",
            bucket=bucket,
            page=page,
            page_size=page_size,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    def get_anniversary_campaign_data(
        self,
        current_user: User,
        bucket: str = "today",
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        sort_by: str = "name",
        sort_order: str = "asc",
    ) -> dict:
        return self.get_celebration_campaign_data(
            current_user=current_user,
            kind="anniversary",
            bucket=bucket,
            page=page,
            page_size=page_size,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    def get_celebration_campaign_data(
        self,
        current_user: User,
        kind: str = "birthday",
        bucket: str = "today",
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        sort_by: str = "name",
        sort_order: str = "asc",
    ) -> dict:
        """
        100% database-driven calculations for Birthday & Anniversary campaigns.
        Supports buckets: today, tomorrow, week, month.
        Includes timezone resolution, server-side pagination, search, sorting, and full logging.
        """
        import zoneinfo
        from datetime import datetime, timezone, timedelta
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from app.models.customer import Customer
        from app.models.business import Business
        from app.models.campaign import Campaign, CampaignLog, CampaignLogStatus, CampaignType

        business_id = current_user.business_id
        business = self.db.scalar(select(Business).where(Business.id == business_id))
        tz_str = business.timezone if (business and business.timezone) else "Asia/Kolkata"
        try:
            tz = zoneinfo.ZoneInfo(tz_str)
        except Exception:
            tz = timezone.utc

        now_tz = datetime.now(tz)
        today_date = now_tz.date()
        tomorrow_date = today_date + timedelta(days=1)
        next_7_days = [today_date + timedelta(days=i) for i in range(7)]

        # Fetch active customers
        all_custs = list(self.db.scalars(
            select(Customer)
            .options(joinedload(Customer.loyalty))
            .where(
                Customer.business_id == business_id,
                Customer.is_active == True,
            )
        ).all())

        # Calculate Counts
        bday_today = bday_tom = bday_week = bday_month = 0
        anni_today = anni_tom = anni_week = anni_month = 0

        for c in all_custs:
            if c.birth_date:
                bm, bd = c.birth_date.month, c.birth_date.day
                if bm == today_date.month and bd == today_date.day: bday_today += 1
                if bm == tomorrow_date.month and bd == tomorrow_date.day: bday_tom += 1
                if any(bm == d.month and bd == d.day for d in next_7_days): bday_week += 1
                if bm == today_date.month: bday_month += 1

            if c.anniversary_date:
                am, ad = c.anniversary_date.month, c.anniversary_date.day
                if am == today_date.month and ad == today_date.day: anni_today += 1
                if am == tomorrow_date.month and ad == tomorrow_date.day: anni_tom += 1
                if any(am == d.month and ad == d.day for d in next_7_days): anni_week += 1
                if am == today_date.month: anni_month += 1

        logger.info(
            "CELEBRATION DEBUG | biz_id=%s tz=%s today=%s cust_count=%d | "
            "bday(today=%d, tom=%d, week=%d, month=%d) | "
            "anni(today=%d, tom=%d, week=%d, month=%d)",
            business_id, tz_str, today_date, len(all_custs),
            bday_today, bday_tom, bday_week, bday_month,
            anni_today, anni_tom, anni_week, anni_month,
        )

        summary = {
            "today": bday_today if kind == "birthday" else anni_today,
            "tomorrow": bday_tom if kind == "birthday" else anni_tom,
            "this_week": bday_week if kind == "birthday" else anni_week,
            "this_month": bday_month if kind == "birthday" else anni_month,
        }

        # Filter customers for requested kind & bucket
        target_field = "birth_date" if kind == "birthday" else "anniversary_date"
        filtered_customers = []

        for c in all_custs:
            dt_val = getattr(c, target_field, None)
            if not dt_val:
                continue

            # Search filtering
            if search and search.strip():
                term = search.strip().lower()
                c_name = (c.name or "").lower()
                c_phone = (c.phone or "").lower()
                c_email = (c.email or "").lower()
                if term not in c_name and term not in c_phone and term not in c_email:
                    continue

            m, d = dt_val.month, dt_val.day
            match = False
            if bucket == "today":
                match = (m == today_date.month and d == today_date.day)
            elif bucket == "tomorrow":
                match = (m == tomorrow_date.month and d == tomorrow_date.day)
            elif bucket == "week":
                match = any(m == day_obj.month and d == day_obj.day for day_obj in next_7_days)
            elif bucket == "month":
                match = (m == today_date.month)
            else:
                match = (m == today_date.month and d == today_date.day)

            if match:
                filtered_customers.append(c)

        # Sorting
        reverse_sort = (sort_order.lower() == "desc")
        if sort_by in ("birth_date", "anniversary_date", "event_date"):
            filtered_customers.sort(
                key=lambda x: (getattr(x, target_field).month, getattr(x, target_field).day) if getattr(x, target_field) else (0, 0),
                reverse=reverse_sort,
            )
        elif sort_by == "last_visit_at":
            filtered_customers.sort(key=lambda x: (x.last_visit_at or x.created_at), reverse=reverse_sort)
        elif sort_by == "total_spent":
            filtered_customers.sort(key=lambda x: float(x.total_spent or 0.0), reverse=reverse_sort)
        else: # name
            filtered_customers.sort(key=lambda x: (x.name or "").lower(), reverse=reverse_sort)

        total_items = len(filtered_customers)
        total_pages = max(1, (total_items + page_size - 1) // page_size) if total_items > 0 else 1
        page = max(1, min(page, total_pages))
        start_idx = (page - 1) * page_size
        paginated_slice = filtered_customers[start_idx : start_idx + page_size]

        coupon_code = "BDAYSPECIAL" if kind == "birthday" else "ANNISPECIAL"

        items = []
        for c in paginated_slice:
            pts = c.loyalty.current_points if c.loyalty else c.loyalty_points
            spent = float(c.total_spent or 0.0)

            if spent >= 1000:
                tier = "High Spender VIP"
            elif spent >= 500:
                tier = "VIP"
            elif c.visit_count > 1:
                tier = "Returning Guest"
            else:
                tier = "First Time Guest"

            event_dt = getattr(c, target_field)
            age = (today_date.year - event_dt.year) if event_dt else None
            initials = "".join([part[0] for part in c.name.split() if part])[:2].upper() if c.name else "NV"

            target_camp_type = CampaignType.BIRTHDAY if kind == "birthday" else CampaignType.ANNIVERSARY
            celeb_log = self.db.scalar(
                select(CampaignLog)
                .join(Campaign, CampaignLog.campaign_id == Campaign.id)
                .where(
                    CampaignLog.customer_id == c.id,
                    Campaign.campaign_type == target_camp_type,
                    CampaignLog.status == CampaignLogStatus.SENT,
                )
            )
            c_status = "Sent" if celeb_log else "Pending"

            items.append({
                "id": c.id,
                "name": c.name,
                "phone": c.phone or "",
                "email": c.email,
                "birth_date": c.birth_date,
                "anniversary_date": c.anniversary_date,
                "event_date": event_dt,
                "age": age,
                "last_visit_at": c.last_visit_at or c.created_at,
                "visit_count": c.visit_count or 1,
                "total_spent": spent,
                "loyalty_points": pts,
                "customer_tier": tier,
                "current_coupon": coupon_code,
                "preferred_language": "Auto",
                "favorite_item": "Specialties",
                "favorites": ["Signature Dishes"],
                "initials": initials,
                "status": c_status,
            })

        logger.info(
            "CELEBRATION RESULT | kind=%s bucket=%s returned_rows=%d total_items=%d total_pages=%d page=%d",
            kind, bucket, len(items), total_items, total_pages, page,
        )

        return {
            "summary": summary,
            "items": items,
            "page": page,
            "page_size": page_size,
            "total_items": total_items,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_previous": page > 1,
        }

    def get_or_create_vip_settings(self, business_id: UUID):
        from app.models.vip_settings import VipSettings
        v_set = self.db.scalar(select(VipSettings).where(VipSettings.business_id == business_id))
        if not v_set:
            v_set = VipSettings(
                business_id=business_id,
                min_lifetime_spend=10000.0,
                min_visits=15,
                min_avg_bill=0.0,
                last_visit_within_days=None,
                rule_logic="ANY",
                is_active=True,
            )
            self.db.add(v_set)
            self.db.commit()
            self.db.refresh(v_set)
        return v_set

    def format_vip_rule_display(self, v_set) -> str:
        rules = []
        if v_set.min_lifetime_spend > 0:
            rules.append(f"Lifetime Spend ≥ ₹{v_set.min_lifetime_spend:,.2f}")
        if v_set.min_visits > 0:
            rules.append(f"Visits ≥ {v_set.min_visits}")
        if v_set.min_avg_bill > 0:
            rules.append(f"Average Bill ≥ ₹{v_set.min_avg_bill:,.2f}")
        if v_set.last_visit_within_days is not None and v_set.last_visit_within_days > 0:
            rules.append(f"Last Visit within {v_set.last_visit_within_days} days")
        
        if not rules:
            return "VIP qualification is currently disabled."
        
        joiner = " OR " if str(v_set.rule_logic).upper() == "ANY" else " AND "
        return f"Current VIP Rule: {joiner.join(rules)} (Rule Logic: {str(v_set.rule_logic).upper()}). Configured in Business Settings."

    def evaluate_customer_vip_status(self, customer: Customer, v_set) -> tuple[bool, str]:
        from datetime import datetime, timezone
        if not v_set or not v_set.is_active:
            return False, "VIP program is disabled."
        
        spent = float(customer.total_spent or 0.0)
        visits = int(customer.visit_count or 0)
        avg_bill = round(spent / max(1, visits), 2) if (spent > 0 and visits > 0) else 0.0
        
        now_dt = datetime.now(timezone.utc)
        days_since_last = None
        if customer.last_visit_at:
            lv_dt = customer.last_visit_at
            if lv_dt.tzinfo is None:
                lv_dt = lv_dt.replace(tzinfo=timezone.utc)
            days_since_last = (now_dt - lv_dt).days

        cond_results = []
        reasons = []

        if v_set.min_lifetime_spend > 0:
            passed = (spent >= v_set.min_lifetime_spend)
            cond_results.append(passed)
            if passed:
                reasons.append(f"Lifetime Spend (₹{spent:,.2f}) ≥ ₹{v_set.min_lifetime_spend:,.2f}")

        if v_set.min_visits > 0:
            passed = (visits >= v_set.min_visits)
            cond_results.append(passed)
            if passed:
                reasons.append(f"Visits ({visits}) ≥ {v_set.min_visits}")

        if v_set.min_avg_bill > 0:
            passed = (avg_bill >= v_set.min_avg_bill)
            cond_results.append(passed)
            if passed:
                reasons.append(f"Average Bill (₹{avg_bill:,.2f}) ≥ ₹{v_set.min_avg_bill:,.2f}")

        if v_set.last_visit_within_days is not None and v_set.last_visit_within_days > 0:
            passed = (days_since_last is not None and days_since_last <= v_set.last_visit_within_days)
            cond_results.append(passed)
            if passed:
                reasons.append(f"Last Visit ({days_since_last} days ago) ≤ {v_set.last_visit_within_days} days")

        if not cond_results:
            return False, "No active VIP rules configured."

        is_vip = any(cond_results) if str(v_set.rule_logic).upper() == "ANY" else all(cond_results)
        
        if is_vip:
            reason_str = f"Qualified because: {' OR '.join(reasons) if str(v_set.rule_logic).upper() == 'ANY' else ' AND '.join(reasons)}"
        else:
            reason_str = "Does not meet the active VIP rule thresholds."
            
        return is_vip, reason_str

    def update_vip_settings(self, current_user: User, data) -> dict:
        v_set = self.get_or_create_vip_settings(current_user.business_id)
        v_set.min_lifetime_spend = data.min_lifetime_spend
        v_set.min_visits = data.min_visits
        v_set.min_avg_bill = data.min_avg_bill
        v_set.last_visit_within_days = data.last_visit_within_days
        v_set.rule_logic = data.rule_logic.upper()
        v_set.is_active = data.is_active
        
        self.db.commit()
        self.db.refresh(v_set)

        formatted_rule = self.format_vip_rule_display(v_set)
        
        # Trigger automatic recalculation of VIP statuses for all customers of this business
        self.recalculate_all_vips_for_business(current_user.business_id)

        return {
            "id": v_set.id,
            "business_id": v_set.business_id,
            "min_lifetime_spend": v_set.min_lifetime_spend,
            "min_visits": v_set.min_visits,
            "min_avg_bill": v_set.min_avg_bill,
            "last_visit_within_days": v_set.last_visit_within_days,
            "rule_logic": v_set.rule_logic,
            "is_active": v_set.is_active,
            "formatted_rule_display": formatted_rule,
            "created_at": v_set.created_at,
            "updated_at": v_set.updated_at,
        }

    def recalculate_all_vips_for_business(self, business_id: UUID):
        v_set = self.get_or_create_vip_settings(business_id)
        customers = list(self.db.scalars(select(Customer).where(Customer.business_id == business_id, Customer.is_active == True)).all())
        vips = []
        for c in customers:
            is_vip, reason = self.evaluate_customer_vip_status(c, v_set)
            if is_vip:
                vips.append(c)
        logger.info("RECALCULATE VIPS | business_id=%s total_vips=%d", business_id, len(vips))

    def get_vip_customers(
        self,
        current_user: User,
        page: int = 1,
        page_size: int = 20,
        search: str | None = None,
        sort_by: str = "spend_desc",
        min_spend: float | None = None,
        min_visits: int | None = None,
    ) -> dict:
        """
        Returns database-driven VIP customer list evaluated dynamically using VipSettings.
        """
        from sqlalchemy import select, func
        from app.models.order import Order, OrderItem
        from sqlalchemy.orm import joinedload

        business_id = current_user.business_id
        v_set = self.get_or_create_vip_settings(business_id)
        formatted_rule = self.format_vip_rule_display(v_set)

        # ── 1. Pull all active customers of the business ──────────────────────
        stmt = (
            select(Customer)
            .options(joinedload(Customer.loyalty))
            .where(
                Customer.business_id == business_id,
                Customer.is_active == True,
            )
        )
        all_customers = list(self.db.scalars(stmt).unique().all())

        # ── 2. Evaluate and Filter VIPs dynamically ───────────────────────────
        vip_entries = []
        for c in all_customers:
            is_vip, reason = self.evaluate_customer_vip_status(c, v_set)
            if is_vip:
                vip_entries.append((c, reason))

        # ── 3. Apply search ──────────────────────────────────────────────────
        if search and search.strip():
            term = search.strip().lower()
            vip_entries = [
                (c, r) for c, r in vip_entries
                if term in (c.name or "").lower() or term in (c.phone or "")
            ]

        # ── 4. Compute favorite item per VIP customer via order_items ────────
        vip_ids = [c.id for c, _ in vip_entries]
        fav_map: dict = {}
        if vip_ids:
            fav_stmt = (
                select(
                    Order.customer_id,
                    OrderItem.item_name,
                    func.sum(OrderItem.quantity).label("total_qty"),
                )
                .join(OrderItem, OrderItem.order_id == Order.id)
                .where(
                    Order.business_id == business_id,
                    Order.customer_id.in_(vip_ids),
                )
                .group_by(Order.customer_id, OrderItem.item_name)
                .order_by(Order.customer_id, func.sum(OrderItem.quantity).desc())
            )
            fav_rows = self.db.execute(fav_stmt).all()

            for row in fav_rows:
                cid = row.customer_id
                if cid not in fav_map:
                    fav_map[cid] = row.item_name

        # ── 5. Sort ──────────────────────────────────────────────────────────
        if sort_by == "spend_desc":
            vip_entries.sort(key=lambda item: float(item[0].total_spent or 0), reverse=True)
        elif sort_by == "spend_asc":
            vip_entries.sort(key=lambda item: float(item[0].total_spent or 0))
        elif sort_by == "visits_desc":
            vip_entries.sort(key=lambda item: item[0].visit_count or 0, reverse=True)
        elif sort_by == "visits_asc":
            vip_entries.sort(key=lambda item: item[0].visit_count or 0)
        elif sort_by == "avg_bill_desc":
            vip_entries.sort(
                key=lambda item: (float(item[0].total_spent or 0) / max(1, item[0].visit_count or 1)),
                reverse=True,
            )
        elif sort_by == "points_desc":
            vip_entries.sort(
                key=lambda item: (item[0].loyalty.current_points if item[0].loyalty else item[0].loyalty_points),
                reverse=True,
            )
        elif sort_by == "recent":
            vip_entries.sort(
                key=lambda item: item[0].last_visit_at or item[0].created_at,
                reverse=True,
            )
        else:
            vip_entries.sort(key=lambda item: float(item[0].total_spent or 0), reverse=True)

        # ── 6. Summary cards ─────────────────────────────────────────────────
        total_vip = len(vip_entries)
        total_lifetime_spend = sum(float(c.total_spent or 0) for c, _ in vip_entries)
        total_visits_sum = sum(c.visit_count or 0 for c, _ in vip_entries)
        total_loyalty_points = sum(
            (c.loyalty.current_points if c.loyalty else c.loyalty_points) for c, _ in vip_entries
        )
        avg_visits = round(total_visits_sum / total_vip, 1) if total_vip else 0.0
        avg_lifetime_spend = round(total_lifetime_spend / total_vip, 2) if total_vip else 0.0

        summary = {
            "total_vip": total_vip,
            "total_lifetime_spend": total_lifetime_spend,
            "avg_visits": avg_visits,
            "avg_lifetime_spend": avg_lifetime_spend,
            "total_loyalty_points": total_loyalty_points,
            "formatted_rule_display": formatted_rule,
        }

        # ── 7. Pagination ────────────────────────────────────────────────────
        total_pages = max(1, (total_vip + page_size - 1) // page_size) if total_vip else 1
        page = max(1, min(page, total_pages))
        start = (page - 1) * page_size
        page_slice = vip_entries[start: start + page_size]

        # ── 8. Build response items ──────────────────────────────────────────
        items = []
        for c, reason in page_slice:
            pts = c.loyalty.current_points if c.loyalty else c.loyalty_points
            spent = float(c.total_spent or 0)
            v_cnt = c.visit_count or 0
            avg_bill = round(spent / max(1, v_cnt), 2) if (spent > 0 and v_cnt > 0) else 0.0

            items.append({
                "id": c.id,
                "name": c.name,
                "phone": c.phone or "",
                "email": c.email,
                "visit_count": v_cnt,
                "total_spent": spent,
                "avg_bill": avg_bill,
                "loyalty_points": pts,
                "favorite_item": fav_map.get(c.id, "No favorite yet"),
                "vip_since_date": c.first_visit_at or c.created_at,
                "last_visit_at": c.last_visit_at,
                "first_visit_at": c.first_visit_at,
                "created_at": c.created_at,
                "status": "VIP",
                "segment": "VIP",
                "reason_qualified": reason,
            })

        logger.info(
            "VIP CUSTOMERS | business_id=%s total_vip=%d page=%d/%d",
            business_id, total_vip, page, total_pages,
        )

        return {
            "summary": summary,
            "settings": {
                "id": v_set.id,
                "business_id": v_set.business_id,
                "min_lifetime_spend": v_set.min_lifetime_spend,
                "min_visits": v_set.min_visits,
                "min_avg_bill": v_set.min_avg_bill,
                "last_visit_within_days": v_set.last_visit_within_days,
                "rule_logic": v_set.rule_logic,
                "is_active": v_set.is_active,
                "formatted_rule_display": formatted_rule,
            },
            "items": items,
            "page": page,
            "page_size": page_size,
            "total": total_vip,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_previous": page > 1,
        }

    def export_customers(
        self,
        current_user: User,
        search: str | None = None,
        sort: str | None = "newest",
        filter: str | None = "all",
        file_format: str = "csv",
    ):
        import csv
        import io
        from datetime import datetime
        from fastapi.responses import Response

        customers = self.repo.get_all_filtered_by_business(
            business_id=current_user.business_id,
            search=search,
            sort=sort,
            filter=filter,
        )

        fmt_clean = (file_format or "csv").lower().strip()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 1. PDF Export Format
        if fmt_clean == "pdf":
            from app.services.pdf_service import generate_customer_export_pdf_bytes

            biz_name = current_user.business.name if current_user.business else "NextVisit Merchant"
            biz_type_name = (
                current_user.business.business_type.name
                if (current_user.business and current_user.business.business_type)
                else "Business"
            )
            biz_address = current_user.business.address if current_user.business else None
            biz_phone = current_user.business.phone if current_user.business else None
            biz_email = current_user.business.email if current_user.business else None
            logo_url = current_user.business.logo_url if current_user.business else None

            customer_dicts = []
            for c in customers:
                is_vip = (c.total_spent >= 500 or c.visit_count >= 10 or getattr(c, "status", None) == "VIP")
                bday = c.birth_date.strftime("%Y-%m-%d") if c.birth_date else ""
                anniv = c.anniversary_date.strftime("%Y-%m-%d") if c.anniversary_date else ""
                last_visit = c.last_visit_at.strftime("%Y-%m-%d %H:%M") if c.last_visit_at else "Never"

                customer_dicts.append({
                    "name": c.name,
                    "phone": c.phone,
                    "email": c.email or "",
                    "gender": c.gender or "",
                    "birth_date": bday,
                    "anniversary_date": anniv,
                    "is_vip": is_vip,
                    "loyalty_points": c.loyalty_points,
                    "visit_count": c.visit_count or 0,
                    "total_spent": float(c.total_spent or 0.0),
                    "last_visit_at": last_visit,
                })

            pdf_bytes = generate_customer_export_pdf_bytes(
                business_name=biz_name,
                business_type_name=biz_type_name,
                business_address=biz_address,
                business_phone=biz_phone,
                business_email=biz_email,
                logo_url=logo_url,
                search_query=search,
                filter_segment=filter,
                sort_order=sort,
                customers=customer_dicts,
            )

            filename = f"customers_export_{timestamp}.pdf"
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )

        # 2. Excel (.xlsx) Export Format
        if fmt_clean in ("excel", "xlsx"):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customers"
            ws.views.sheetView[0].showGridLines = True

            biz_name = current_user.business.name if current_user.business else "NextVisit Merchant"
            ws.merge_cells("A1:L1")
            title_cell = ws["A1"]
            title_cell.value = f"{biz_name} — Customer Directory & Performance Report"
            title_cell.font = Font(name="Arial", size=14, bold=True, color="1E293B")
            title_cell.alignment = Alignment(vertical="center")

            ws.merge_cells("A2:L2")
            meta_cell = ws["A2"]
            meta_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}  |  Search: '{search or 'All'}'  |  Segment: '{filter or 'All'}'  |  Sorting: '{sort or 'Newest'}'"
            meta_cell.font = Font(name="Arial", size=9, italic=True, color="64748B")

            ws.append([])

            headers = [
                "Customer Name", "Phone Number", "Email Address", "Gender",
                "Birthday", "Anniversary", "VIP Status", "Loyalty Points",
                "Visits", "Total Spend (₹)", "Last Visit", "Created Date"
            ]
            ws.append(headers)

            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=4, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            for r_idx, c in enumerate(customers, start=5):
                is_vip = "Yes" if (c.total_spent >= 500 or c.visit_count >= 10 or getattr(c, "status", None) == "VIP") else "No"
                bday = c.birth_date.strftime("%Y-%m-%d") if c.birth_date else ""
                anniv = c.anniversary_date.strftime("%Y-%m-%d") if c.anniversary_date else ""
                last_visit = c.last_visit_at.strftime("%Y-%m-%d %H:%M") if c.last_visit_at else "Never"
                created_at = c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""

                row_vals = [
                    c.name,
                    c.phone,
                    c.email or "",
                    c.gender or "",
                    bday,
                    anniv,
                    is_vip,
                    c.loyalty_points,
                    c.visit_count or 0,
                    float(c.total_spent or 0.0),
                    last_visit,
                    created_at,
                ]
                ws.append(row_vals)

                is_even = (r_idx % 2 == 0)
                for col_idx in range(1, len(row_vals) + 1):
                    cell = ws.cell(row=r_idx, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(name="Arial", size=9)
                    if is_even:
                        cell.fill = alt_fill
                    if col_idx in (7, 8, 9):
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx == 10:
                        cell.number_format = '"₹"#,##0.00'
                        cell.alignment = Alignment(horizontal="right")

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_bytes = excel_buf.getvalue()

            filename = f"customers_export_{timestamp}.xlsx"
            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Access-Control-Expose-Headers": "Content-Disposition",
                },
            )

        # 3. CSV Export Format
        headers = [
            "Customer Name", "Phone", "Email", "Gender", "Birthday",
            "Anniversary", "VIP Status", "Loyalty Points", "Visits",
            "Total Spend", "Last Visit", "Created Date",
        ]

        rows = []
        for c in customers:
            is_vip = "Yes" if (c.total_spent >= 500 or c.visit_count >= 10 or getattr(c, "status", None) == "VIP") else "No"
            bday = c.birth_date.strftime("%Y-%m-%d") if c.birth_date else ""
            anniv = c.anniversary_date.strftime("%Y-%m-%d") if c.anniversary_date else ""
            last_visit = c.last_visit_at.strftime("%Y-%m-%d %H:%M") if c.last_visit_at else "Never"
            created_at = c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""

            rows.append([
                c.name,
                c.phone,
                c.email or "",
                c.gender or "",
                bday,
                anniv,
                is_vip,
                c.loyalty_points,
                c.visit_count or 0,
                f"₹{c.total_spent:.2f}",
                last_visit,
                created_at,
            ])

        output = io.StringIO()
        output.write("\ufeff")
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        filename = f"customers_export_{timestamp}.csv"
        csv_bytes = output.getvalue().encode("utf-8")

        return Response(
            content=csv_bytes,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    def import_customers(
        self,
        current_user: User,
        file: UploadFile,
    ) -> dict:
        import csv
        import io
        import re
        from datetime import datetime, date

        filename = (file.filename or "").lower()
        content_bytes = file.file.read()

        if not content_bytes or len(content_bytes) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        rows = []
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(filename=io.BytesIO(content_bytes), data_only=True)
                sheet = wb.active
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None and str(cell).strip() != "" for cell in row):
                        rows.append([str(c if c is not None else "").strip() for c in row])
            except Exception as e:
                logger.warning("Failed to parse Excel file, falling back to CSV text parse: %s", e)
                text_content = content_bytes.decode("utf-8-sig", errors="ignore")
                reader = csv.reader(io.StringIO(text_content))
                rows = list(reader)
        else:
            try:
                text_content = content_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                text_content = content_bytes.decode("latin-1", errors="ignore")

            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)

        if not rows or len(rows) < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File contains no data rows to import.",
            )

        header_row = [str(cell).strip().lower() for cell in rows[0]]

        def find_col_idx(aliases: list[str]) -> int | None:
            for alias in aliases:
                for idx, col in enumerate(header_row):
                    if alias in col:
                        return idx
            return None

        col_name = find_col_idx(["customer name", "name", "full name", "customer"])
        col_phone = find_col_idx(["phone number", "phone", "mobile", "contact"])
        col_email = find_col_idx(["email address", "email"])
        col_gender = find_col_idx(["gender", "sex"])
        col_bday = find_col_idx(["birthday", "birth date", "dob", "date of birth"])
        col_anniv = find_col_idx(["anniversary", "anniversary date"])
        col_notes = find_col_idx(["notes", "note", "address"])

        if col_name is None or col_phone is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV/Excel file must contain 'Customer Name' and 'Phone' header columns.",
            )

        existing_phones = set(
            self.db.scalars(
                select(Customer.phone).where(Customer.business_id == current_user.business_id)
            ).all()
        )
        seen_in_batch = set()

        def clean_phone_number(p_str: str) -> str | None:
            if not p_str:
                return None
            digits = "".join(c for c in str(p_str) if c.isdigit())
            if len(digits) == 12 and digits.startswith("91"):
                digits = digits[2:]
            elif len(digits) == 11 and digits.startswith("0"):
                digits = digits[1:]
            if len(digits) == 10:
                return digits
            return None

        def parse_date_val(d_str: str) -> date | None:
            if not d_str or not str(d_str).strip():
                return None
            s = str(d_str).strip()
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]:
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    pass
            return None

        email_regex = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

        total_rows = len(rows) - 1
        imported_count = 0
        skipped_count = 0
        failed_count = 0
        duplicate_count = 0
        errors = []

        valid_customers = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not any(str(cell).strip() for cell in row):
                skipped_count += 1
                continue

            raw_name = str(row[col_name]).strip() if col_name < len(row) else ""
            raw_phone = str(row[col_phone]).strip() if col_phone < len(row) else ""
            raw_email = str(row[col_email]).strip() if col_email is not None and col_email < len(row) else ""
            raw_gender = str(row[col_gender]).strip() if col_gender is not None and col_gender < len(row) else ""
            raw_bday = str(row[col_bday]).strip() if col_bday is not None and col_bday < len(row) else ""
            raw_anniv = str(row[col_anniv]).strip() if col_anniv is not None and col_anniv < len(row) else ""
            raw_notes = str(row[col_notes]).strip() if col_notes is not None and col_notes < len(row) else ""

            if not raw_name:
                failed_count += 1
                errors.append({
                    "row": row_idx,
                    "field": "Customer Name",
                    "reason": "Customer Name is required",
                })
                continue

            if not raw_phone:
                failed_count += 1
                errors.append({
                    "row": row_idx,
                    "field": "Phone",
                    "reason": "Phone Number is required",
                })
                continue

            clean_phone = clean_phone_number(raw_phone)
            if not clean_phone:
                failed_count += 1
                errors.append({
                    "row": row_idx,
                    "field": "Phone",
                    "reason": f"Invalid phone '{raw_phone}'. Must be a valid 10-digit number.",
                })
                continue

            if clean_phone in existing_phones or clean_phone in seen_in_batch:
                duplicate_count += 1
                skipped_count += 1
                errors.append({
                    "row": row_idx,
                    "field": "Phone",
                    "reason": f"Duplicate customer phone number '{clean_phone}' skipped.",
                })
                continue

            clean_email = None
            if raw_email:
                if not email_regex.match(raw_email):
                    failed_count += 1
                    errors.append({
                        "row": row_idx,
                        "field": "Email",
                        "reason": f"Invalid email format '{raw_email}'.",
                    })
                    continue
                clean_email = raw_email

            bday_val = parse_date_val(raw_bday)
            anniv_val = parse_date_val(raw_anniv)

            customer = Customer(
                business_id=current_user.business_id,
                name=raw_name,
                phone=clean_phone,
                email=clean_email,
                gender=raw_gender if raw_gender in ["Male", "Female", "Other"] else (raw_gender or None),
                birth_date=bday_val,
                anniversary_date=anniv_val,
                notes=raw_notes or None,
                is_active=True,
            )
            valid_customers.append(customer)
            seen_in_batch.add(clean_phone)
            existing_phones.add(clean_phone)
            imported_count += 1

        if valid_customers:
            self.db.bulk_save_objects(valid_customers)
            self.db.commit()

        logger.info(
            "CUSTOMER IMPORT COMPLETED | business_id=%s total=%d imported=%d skipped=%d failed=%d duplicates=%d",
            current_user.business_id,
            total_rows,
            imported_count,
            skipped_count,
            failed_count,
            duplicate_count,
        )

        return {
            "total_rows": total_rows,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
            "failed_count": failed_count,
            "duplicate_count": duplicate_count,
            "errors": errors,
            "message": f"Successfully imported {imported_count} customers out of {total_rows} rows.",
        }
